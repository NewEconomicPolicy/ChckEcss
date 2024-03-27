# -------------------------------------------------------------------------------
# Name:        generate_charts_funcs.py.py
# Purpose:     examine two directories each with Ecosse output files and compare differences
# Author:      Mike Martin
# Created:     07/03/2015
# Licence:     <your licence>
# -------------------------------------------------------------------------------

__prog__ = 'generate_charts_funcs.py'
__version__ = '0.0.1'

# Version history
# ---------------
# 0.0.1  Wrote.
#

from os.path import join, isdir, split, isfile, exists
from os import remove
import csv

from PyQt5.QtWidgets import QApplication

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from analyse_ecosse_output import format_out_files

METRICS_GROUPS = {'carbon': list(['BIOC', 'CO2', 'DPMC', 'HUMC', 'RPMC', 'TOTC']),
                  'nitrogen': list(['BION', 'DPMN', 'HUMN', 'RPMN', 'TOTN', 'SOILW']),
                  'balance_n': list(['NitN2O', 'PNitN2O', 'DenN2O']),
                  'nitrate': list(['NO3N', 'NON', 'N2ON', 'NITRIFN', 'MINERN', 'LEACHN', 'NH4N']),
                  'npp': list(['NPP', 'CH4'])}

PRFRD_LINE_WDTH = 25000  # 100020 taken from chart_example.py     preferred line width in EMUs
FNAME_LINK_STR = '_vs_'
ERROR_STR = '*** Error *** '

METRIC_MPPNGS_V6_2 = {'BIOC': 'bio_c', 'CO2': 'co2_c', 'DPMC': 'dpm_c', 'HUMC': 'hum_c', 'RPMC': 'rpm_c',
                      'TOTC': 'total_soc',
                      'BION': 'bio_n', 'DPMN': 'dpm_n', 'HUMN': 'hum_n', 'RPMN': 'rpm_n', 'TOTN': 'total_son',
                      'SOILW': 'avail_water',
                      'NON': 'no_n', 'N2ON': 'n2o_n',
                      'NO3N': 'no3_n', 'NITRIFN': 'nitrification_n', 'MINERN': 'net_mineralised_n',
                      'LEACHN': 'leached_n',
                      'NH4N': 'nh4_n', 'NPP': 'npp_adj', 'CH4': 'ch4_c'}

METRIC_MPPNGS_V6_3 = {'BIOC': 'BIO_C', 'CO2': 'CO2_C', 'DPMC': 'DPM_C', 'HUMC': 'HUM_C', 'RPMC': 'RPM_C',
                      'TOTC': 'Total_SOC',
                      'BION': 'BIO_N', 'DPMN': 'DPM_N', 'HUMN': 'HUM_N', 'RPMN': 'RPM_N', 'TOTN': 'Total_SON',
                      'SOILW': 'Avail_Water',
                      'NON': 'NO_N', 'N2ON': 'N2O_N',
                      'NO3N': 'NO3_N', 'NITRIFN': 'Nitrif_N', 'MINERN': 'Mineralised_N',
                      'LEACHN': 'Leached_N',
                      'NH4N': 'NH4_N', 'NPP': 'NPP_ADJ', 'CH4': 'CH4_C'}
SKIP_NPP = True

def _search_for_metric_and_map(group, summary_metrics, metric):
    """
    locate metric is summary.out metrics
    """
    mtrc_mppd = METRIC_MPPNGS_V6_2[metric]
    if mtrc_mppd not in summary_metrics:
        mtrc_mppd = METRIC_MPPNGS_V6_3[metric]
        if mtrc_mppd not in summary_metrics:
            mtrc_mppd = None
            mess = ' in SUMMARY.OUT from ECOSSE version 6.2 or 6.3'
            print(ERROR_STR + 'group: ' + group + ' no mapping for metric: ' + metric + mess)
            QApplication.processEvents()

    return mtrc_mppd

def generate_charts(form):
    """
    C
    """

    # gather directories from the GUI
    # ===============================
    ref_dir = form.w_lbl03.text()
    ref_id = form.w_ref_id.text()
    if ref_id == '':
        print('Reference identifier cannot be blank')
        return

    compare_fname = ref_id

    targ1_dir = form.w_lbl04.text()
    targ1_id = form.w_targ1_id.text()
    if targ1_id == '':
        print('Target 1 identifier cannot be blank')
        return

    idents = [ref_id, targ1_id]
    compare_fname += '_' + targ1_id

    if form.w_targ2_also.isChecked():
        targ2_also_flag = True
        targ2_dir = form.w_lbl06.text()
        targ2_id = form.w_targ2_id.text()
        if targ2_id == '':
            print('Target 2 identifier cannot be blank')
            return

        idents.append(targ2_id)
        compare_fname += '_' + targ2_id

        max_col_val = 3
    else:
        max_col_val = 2
        targ2_also_flag = False
        targ2_dir = None

    rslts_dir = form.w_lbl13.text()
    water_dep_str = form.w_water_dep.text()
    water_dep = float(water_dep_str)

    # trap possible error
    # ===================
    for dir_name in list([ref_dir, targ1_dir, rslts_dir]):
        if isdir(dir_name):
            continue
        else:
            print(dir_name + ' does not exist')
            return

    sim_dir_names = {'ref': ref_dir, 'targ1': targ1_dir}
    if targ2_also_flag:
        sim_dir_names['targ2'] = targ2_dir

    # Excel file name of comparisons
    # ==============================
    compare_fname += '.xlsx'
    charts_fname = join(rslts_dir, compare_fname)
    if exists(charts_fname):
        try:
            remove(charts_fname)
        except (OSError, IOError) as err:
            print('Failed to delete output file {}'.format(err))
            return -1

    wrkbk = Workbook()
    header_rec = list(idents)

    # create and record sheets for charts
    # ===================================
    save_sheets_flag = True
    wrkshts_group = {}
    for group in METRICS_GROUPS:

        # set up worksheet for each group
        # ===============================
        if group == 'carbon':
            wrksht = wrkbk.active
        else:
            wrksht = wrkbk.create_sheet()
        wrksht.title = group
        wrkshts_group[group] = wrksht

    # retrieve data from .OUT files and write to sheets
    # =================================================
    balance_set = None
    for group in METRICS_GROUPS:
        metric_group = METRICS_GROUPS[group]

        # retrieve the last column for each metric - the sum of all layers
        # ================================================================
        results = {}
        max_num_pts = 999999999
        for sim_name in sim_dir_names:
            print()
            dir_name = sim_dir_names[sim_name]
            results[sim_name] = {}

            # use SUMMARY.OUT if it exists
            # ============================
            summary_set = None
            summary_metrics = None
            nout_files = format_out_files(form, label_string_flag=False)
            if form.w_smmry_only.isChecked() or nout_files < 30:
                summary_out = join(dir_name, 'SUMMARY.OUT')
                if isfile(summary_out):
                    summary_set = _get_out_file_contents(dir_name)
                    summary_metrics = list(summary_set.keys())
                else:
                    print(ERROR_STR + summary_out + ' is not a file')
                    continue

            if group == 'balance_n':
                balance_set = _get_out_file_contents(dir_name, group.upper() + '.OUT')
                if balance_set is None:
                    metric_group = []

            for metric in metric_group:
                if metric == 'NPP' and SKIP_NPP:
                    continue

                if group == 'balance_n':
                    result = balance_set[metric]
                else:
                    if summary_set is None:
                        result = _read_last_column(dir_name, metric, water_dep)
                    else:
                        mtrc_mppd = _search_for_metric_and_map(group, summary_metrics, metric)
                        if mtrc_mppd is None:
                            continue
                        try:
                            result = summary_set[mtrc_mppd]
                        except KeyError as err:
                            print(ERROR_STR + 'metric ' + str(err) + ' not found')
                            continue

                results[sim_name][metric] = result
                npts = len(results[sim_name][metric])
                max_num_pts = min(npts, max_num_pts)
                print('Read {} lines from run {}\tmetric: {}'.format(len(result), sim_name, metric))

        if max_num_pts == 0:
            print('Nothing to plot for ' + group + ' group - max_num_pts = 0')
        #    continue
        # else:
        #    save_sheets_flag = True

        # Excel section - write charts to appropriate group
        # =================================================
        nrow_chart = 2
        for ic, metric in enumerate(metric_group):
            if metric not in results['ref']:
                print('Will skip metric ' + metric + ' - not present in results')
                continue

            wrksht = wrkbk.create_sheet()
            wrksht.title = metric

            rows = list([header_rec])

            for ic, ref_val in enumerate(results['ref'][metric]):
                if ic >= max_num_pts:
                    break

                try:
                    row_rec = [float(ref_val)]
                except ValueError as e:
                    print(
                        'ValueError converting ref_val: ' + ref_val + '\tmetric: ' + metric + '\tsim name: ' + sim_name)
                    break

                targ1_val = results['targ1'][metric][ic]
                row_rec.append(float(targ1_val))

                if targ2_also_flag:
                    targ2_val = results['targ2'][metric][ic]
                    row_rec.append(float(targ2_val))

                rows.append(row_rec)

            for row in rows:
                wrksht.append(row)

            metric_chart = LineChart()
            metric_chart.style = 13
            if metric == 'SOILW':
                metric_chart.title = metric + '\tdepth to bottom of SOM layer (cms): ' + water_dep_str
                metric_chart.y_axis.title = 'mm'
            else:
                metric_chart.title = metric
                metric_chart.y_axis.title = 'kgC/ha'
            metric_chart.x_axis.title = 'Time step'
            metric_chart.height = 10
            metric_chart.width = 20

            nrows = len(rows)
            data = Reference(wrksht, min_col=1, min_row=1, max_col=max_col_val, max_row=nrows)
            metric_chart.add_data(data, titles_from_data=True)

            # Style the lines
            # ===============
            sref = metric_chart.series[0]
            sref.graphicalProperties.line.width = PRFRD_LINE_WDTH
            sref.smooth = True

            targ1 = metric_chart.series[1]
            targ1.graphicalProperties.line.solidFill = "FF0000"
            targ1.graphicalProperties.line.width = PRFRD_LINE_WDTH
            targ1.smooth = True

            if targ2_also_flag:
                targ2 = metric_chart.series[2]
                targ2.graphicalProperties.line.solidFill = "00AAAA"
                targ2.graphicalProperties.line.width = PRFRD_LINE_WDTH
                targ2.smooth = True  # Make the line smooth

            # now write to previously created sheet
            # =====================================
            wrkshts_group[group].add_chart(metric_chart, "A" + str(nrow_chart))
            nrow_chart += 20

    if save_sheets_flag:
        try:
            wrkbk.save(charts_fname)
            print('Created: ' + charts_fname)
        except PermissionError as e:
            print(str(e) + ' - could not create: ' + charts_fname)

    return

def _read_last_column(inp_dir, var_name, water_dep):
    # check file exists and read all lines
    # ====================================
    readings = []
    fname = var_name + '.OUT'
    inp_fname = join(inp_dir, fname)
    if not isfile(inp_fname):
        print('File ' + inp_fname + ' does not exist')
        return readings

    fobj = open(inp_fname, 'r')
    contents = fobj.read()
    content_lines = contents.splitlines()
    fobj.close()

    # step through each line
    # ======================
    if var_name == 'TOTC' or var_name == 'TOTN':
        nhead_lines = 5
    else:
        nhead_lines = 2

    # get soil layer depth - maximum depth is 300cms
    # ====================
    if var_name == 'SOILW':
        ncols = int(water_dep / 5.0)
        ncols = min(60, max(1, ncols))
        end_col = ncols + 3

    # main loop
    # =========
    read_next_rec_flag = True
    for nline, rec in enumerate(content_lines):
        if nline < nhead_lines:
            continue

        rec_atoms = rec.split()

        # skip later version additional imbedded header records for TOTC only
        # this stanza is ignored in output from earlier versions
        if var_name == 'TOTC':
            if rec_atoms[0] == 'Inert' and rec_atoms[1] == 'Organic':
                read_next_rec_flag = False

            elif rec_atoms[0] == 'Total' and rec_atoms[1] == 'soil':
                read_next_rec_flag = True

            elif read_next_rec_flag:
                readings.append(rec_atoms[-1])
                read_next_rec_flag = True

        elif var_name == 'SOILW':
            if rec_atoms[0] == 'Available' and rec_atoms[2] == 'at':
                read_next_rec_flag = False

            elif rec_atoms[0] == 'Available' and rec_atoms[2] == '(mm)':
                read_next_rec_flag = True

            elif read_next_rec_flag:
                avail_water = 0.0
                for val in rec_atoms[3:end_col]:
                    avail_water += float(val)
                readings.append(avail_water)
                read_next_rec_flag = True

        else:
            # neither SOILW or TOTC
            # =====================
            readings.append(rec_atoms[-1])

    return readings


def _read_site_file(inp_dir):
    # check file exists and read all lines
    # ====================================
    readings = []
    fname = 'SITE.TXT'
    inp_fname = join(inp_dir, fname)
    if not isfile(inp_fname):
        print('File ' + inp_fname + ' does not exist')
        return readings

    fobj = open(inp_fname, 'r')
    contents = fobj.read()
    content_lines = contents.splitlines()
    fobj.close()
    soil_depth = 0
    return soil_depth


class EcosseDialect(csv.excel):
    """Dialect class for reading in ECOSSE output files with the csv module."""
    delimiter = ' '
    skipinitialspace = True


def _get_out_file_contents(folder, outfile='SUMMARY.OUT'):
    """
    reads ECOSSE summary results file as a dictionary.

    Arguments:
    folder name for this grid cell's simulation outputs

    Returns:
    Pandas dataframe of the results with date index
    """
    func_name = __prog__ + ' _get_out_file_contents'

    path = join(folder, outfile)
    if not isfile(path):
        print('File {} does not exist - function {}'.format(path, func_name))
        return None

    summary = {}
    nline = 0
    with open(path, 'r') as f:
        reader = csv.reader(f, dialect=EcosseDialect)
        if outfile == 'SUMMARY.OUT':
            nline += 1
            next(reader)  # Skip the units description line

        try:
            columns = next(reader)
        except StopIteration as e:
            print('Error {} on line {} reading {} will skip'.format(e, nline + 1, outfile))
            return None

        for column in columns:
            summary[column] = []

        # step through each row
        # =====================
        nbad_values = 0
        max_bad_values = 10
        for irow, row in enumerate(reader):
            for icol, val in enumerate(row):
                try:
                    float(val)
                except ValueError as e:
                    nbad_values += 1
                    if nbad_values < max_bad_values:
                        print('Error {} on line {} will skip'.format(e, irow + 1))
                    val = -999.0

                summary[columns[icol]].append(float(val))

        print('Finished reading {} with {} bad values\n'.format(path, nbad_values))
        QApplication.processEvents()

    return summary
